from datetime import date
from openpyxl import *
import smtplib

# creates SMTP session
s = smtplib.SMTP('smtp.gmail.com', 587)

# start TLS for security
s.starttls()

# Authentication
##### At first enable two step verification of email id and then go to app and generate password 

s.login("xyz@gmail.com", "put your app password here after two step verification")

def stringcheck(val):
    #print(val)
    today = date.today()

    # dd/mm/YY
    d1 = today.strftime("%d")
    m1 = today.strftime("%m")
    y1 = today.strftime("%Y")

    string1 = int(val[0:2])
    string2 = int(val[3:5])
    string3 = int(val[6:10])


    res1 = int(d1) - string1
    res2 = int(m1) - string2
    res3 = int(y1) - string3

    if(res1 == 0) and (res2 == 0) and (res3 == 0):
        #print("OK")
        message = """\
Subject: Device Reminder, Siemens Goa

Please return the device within today.

Serial Number  -  """

        message2 = """\

        Expected Return Date  -  """

                # sending the mail
        #s.sendmail("siemensgoa2022@gmail.com", email_cell.value, message + str(serial_cell.value)  + message2 + str(retdate_cell.value))

                # terminating the session
        #s.quit()
    else:
        print("fg")

    
    


wb = load_workbook('tkinterfinals.xlsx')
# create the sheet object 

sh = wb.active
for i in range(1, sh.max_row+1):
    serial_cell = sh.cell(row=i, column=1)
    email_cell = sh.cell(row=i, column=4)
    retdate_cell = sh.cell(row=i, column=5)
    statav_cell = sh.cell(row=i, column=7)
    if(statav_cell.value == "Not Available"):
        stringcheck(retdate_cell.value)
                
s.quit()


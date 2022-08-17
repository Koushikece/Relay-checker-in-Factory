# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
from tkinter import ttk, Tk, Toplevel
import smtplib
  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
wb = load_workbook('tkinterfinals.xlsx') 
  
# create the sheet object 
sheet = wb.active 
  
window = Tk()
window.rowconfigure(0, weight=1)
window.columnconfigure(0, weight=1)
window.state('zoomed')


# creates SMTP session
s = smtplib.SMTP('smtp.gmail.com', 587)

# start TLS for security
s.starttls()

# Authentication

s.login("xyz@gmail.com", "put your app password after two step verification")



def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    
    # write given data to an excel spreadsheet 
    # at particular location
    sheet.cell(row=1, column=1).value = "Serial No"
    sheet.cell(row=1, column=2).value = "MLFB No"
    sheet.cell(row=1, column=3).value = "Name"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Return_Date"
    sheet.cell(row=1, column=6).value = "Phone Number"
    sheet.cell(row=1, column=7).value = "Status"
    
  
  
# Function to set focus (cursor)

    
def focus1(event): 
    # set focus on the course_field box 
    course_field.focus_set() 
  
  
# Function to set focus 
def focus2(event): 
    # set focus on the sem_field box 
    sem_field.focus_set() 
  
  
# Function to set focus 
def focus3(event): 
    # set focus on the form_no_field box 
    form_no_field.focus_set() 

  
# Function to set focus 
def focus4(event): 
    # set focus on the contact_no_field box 
    contact_no_field.focus_set() 

  
  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box
    #serial_field.delete(0, END)
    Serial_field.delete(0, END)
    name_field.delete(0, END) 
    course_field.delete(0, END) 
    sem_field.delete(0, END) 
    form_no_field.delete(0, END) 
    '''contact_no_field.delete(0, END) 
    email_id_field.delete(0, END) 
    address_field.delete(0, END)'''
  
  
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
    sh = wb.active
    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == ""):
              
        print("empty input")
        
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable
        for i in range(1, sh.max_row+1):
            cell_val5 = sh.cell(row=i, column=1)
            cell_mlfb = sh.cell(row=i, column=2)
            namec = sh.cell(row=i, column=3)
            retc = sh.cell(row=i, column=5)
        
            if(cell_val5.value == Serial_field1.get()):
                current_row = i-1
                sheet.cell(row=current_row + 1, column=3).value = name_field.get() 
                sheet.cell(row=current_row + 1, column=4).value = course_field.get() 
                sheet.cell(row=current_row + 1, column=5).value = sem_field.get() 
                sheet.cell(row=current_row + 1, column=6).value = form_no_field.get()
                sheet.cell(row=current_row + 1, column=7).value = "Not Available"
                #break
#--------------------------------mail
                message = """\
Subject: Device Status, Siemens Goa
Someone has taken a device.

Serial Number  -  """
                message1 = """\

        Name  -  """
                message2 = """\
        Expected Return Date  -  """


                # sending the mail to the specific mail id
                s.sendmail("xyz@gmail.com", "reciever@gmail.com", message + str(cell_val5.value) + message1 + str(namec.value) + message2 + str(retc.value))
                #s.sendmail("siemensgoa2022@gmail.com", "kanchankaity1396@gmail.com", message + str(cell_val5.value) + message2 + str(retc.value))

                # terminating the session
                #s.quit()
                
#------------------------program
                
                for widget in page8.winfo_children():
                    widget.destroy()
                show_frame(page8)

                page8.config(background='light green')
                page8_label = Label(page8, text='Submitted Successfully ', font=('Arial', 20, 'bold') , bg="light green")
                page8_label.place(x=570, y=200)

                back4 = Button(page8, text="Back", fg="Black", 
                                            bg="Red", command=lambda: show_frame(page4)) 
                back4.place(x = 50, y = 500 ,width = 100 , height = 25)

                name14 = Label(page8, text="Serial Number", font=('Arial' , 12, 'bold') , bg="light green")
                name14.place(x=200, y=400)
                name14 = Label(page8, text=cell_val5.value, font=('Arial' , 12, 'bold') , bg="light green")
                name14.place(x=350, y=400)

                name14 = Label(page8, text="MLFB Number", font=('Arial' , 12, 'bold') , bg="light green")
                name14.place(x=950, y=400)
                name14 = Label(page8, text=cell_mlfb.value, font=('Arial' , 12, 'bold') , bg="light green")
                name14.place(x=1100, y=400)
                
        

        Serial_field1.delete(0, END)
  
        # save the file 
        wb.save('tkinterfinals.xlsx')

        
  
        # set focus on the name_field box 
        name_field.focus_set() 
  
        # call the clear() function 
        clear() 

def insert_page3():
    if (Serial_field.get() == "" ):
        print("empty input")
    else:
        current_row = sheet.max_row 
        current_column = sheet.max_column 

        sheet.cell(row=current_row + 1, column=1).value = Serial_field.get()
        sheet.cell(row=current_row + 1, column=2).value = Serial_field4.get()
        sheet.cell(row=current_row + 1, column=7).value = "Available"
        wb.save('tkinterfinals.xlsx') 
  
  
        # call the clear() function 
        clear()
        Serial_field4.delete(0, END)
        show_frame(page9)

#sh = wb.active
def check():
    sh = wb.active
    
    for i in range(1, sh.max_row+1):
        cell_obj_av = sh.cell(row=i, column=7)
        cell_obj = sh.cell(row=i, column=1)
        mlfb_cell = sh.cell(row=i, column=2)
        name_cell = sh.cell(row=i, column=3)
        email_cell = sh.cell(row=i, column=4)
        date_cell = sh.cell(row=i, column=5)
        phone_cell = sh.cell(row=i, column=6)
        

    
        if(cell_obj_av.value == "Available") and (cell_obj.value == Serial_field1.get()):
            show_frame(page2)
            
        if(cell_obj_av.value == "Not Available") and (cell_obj.value == Serial_field1.get()):
            name_cell = sh.cell(row=i, column=3)
            name_cell1 = name_cell.value

            for widget in page5.winfo_children():
                widget.destroy()
    
            show_frame(page5)

            page5.config(background='light green')
            page5_label = Label(page5, text='NOT AVAILABLE', font=('Arial', 20, 'bold') , bg="light green")
            page5_label.place(x=570, y=50)
            heading = Label(page5, text="Form", bg="light green")

            mlfb_dis = Label(page5, text="MLFB Number", font=('Arial' , 15, 'bold') , bg="light green")
            mlfb_dis.place(x=50, y=200)

            name = Label(page5, text="Name", font=('Arial' , 15, 'bold') , bg="light green")
            name.place(x=335, y=200)


            email = Label(page5, text="Email", font=('Arial' , 15, 'bold') , bg="light green")
            email.place(x=570, y=200)


            date = Label(page5, text="Return Date", font=('Arial' , 15, 'bold') , bg="light green")
            date.place(x=850, y=200)


            numb = Label(page5, text="Phone Number", font=('Arial' , 15, 'bold') , bg="light green")
            numb.place(x=1060, y=200)

            mlfb_dis1 = Label(page5, text=mlfb_cell.value, font=('Arial' , 12, 'bold') , bg="light green")
            mlfb_dis1.place(x=20, y=300)
            name1 = Label(page5, text=name_cell.value, font=('Arial' , 12, 'bold') , bg="light green")
            name1.place(x=320, y=300)
            email = Label(page5, text=email_cell.value, font=('Arial' , 12, 'bold') , bg="light green")
            email.place(x=540, y=300)
            date1 = Label(page5, text=date_cell.value, font=('Arial' , 12, 'bold') , bg="light green")
            date1.place(x=860, y=300)
            numb1 = Label(page5, text=phone_cell.value, font=('Arial' , 12, 'bold') , bg="light green")
            numb1.place(x=1070, y=300)

            back3 = Button(page5, text="Back", fg="Black", 
                            bg="Red", command=lambda: show_frame(page4)) 
            back3.place(x = 50, y = 500 ,width = 100 , height = 25)

            Serial_field1.delete(0, END)

            clear()
    
            
def returnitem():
    sh = wb.active

    for widget in page7.winfo_children():
         widget.destroy()
    ind = 0
    for i in range(1, sh.max_row+1):
        cell_val2 = sh.cell(row=i, column=1)
        stat2 = sh.cell(row=i, column=7)
        if(cell_val2.value == Serial_field2.get()):
            current_row = i-1
            if(stat2.value == "Not Available"):
                sheet.cell(row=current_row + 1, column=7).value = "Available"
                wb.save('tkinterfinals.xlsx')
                #show_frame(page7)
                show_frame(page7)

                page7.config(background='light green')
                page7_label = Label(page7, text='Returned Successfully ', font=('Arial', 20, 'bold') , bg="light green")
                page7_label.place(x=570, y=200)

                back3 = Button(page7, text="Back", fg="Black", 
                                            bg="Red", command=lambda: show_frame(page6)) 
                back3.place(x = 50, y = 500 ,width = 100 , height = 25)
                
                name12 = Label(page7, text="Serial Number", font=('Arial' , 15, 'bold') , bg="light green")
                name12.place(x=460, y=400)
                name12 = Label(page7, text=cell_val2.value, font=('Arial' , 15, 'bold') , bg="light green")
                name12.place(x=650, y=400)
                Serial_field2.delete(0, END)

                ind = 1
                break
            else:
                show_frame(page11)
                page11.config(background='light green')
                page11_label = Label(page11, text='Item is already available ', font=('Arial', 20, 'bold') , bg="light green")
                page11_label.place(x=550, y=200)
                back11 = Button(page11, text="Back", fg="Black", 
                                bg="Red", command=lambda: show_frame(page6)) 
                back11.place(x = 50, y = 500 ,width = 100 , height = 25)

                ind = 1
    if(ind!=1):
        show_frame(page12)
        page12.config(background='light green')
        page12_label = Label(page12, text='Please Enter Correct Item ', font=('Arial', 20, 'bold') , bg="light green")
        page12_label.place(x=570, y=200)
        back12 = Button(page12, text="Back", fg="Black", 
                        bg="Red", command=lambda: show_frame(page6)) 
        back12.place(x = 50, y = 500 ,width = 100 , height = 25)


def scroll_list():
    sh = wb.active
    for widget in page10.winfo_children():
         widget.destroy()

    show_frame(page10)

    page10.config(background='light green')
    page10_label = Label(page10, text='ITEM LIST ', font=('Arial', 20, 'bold') , bg="light green")
    page10_label.place(x=600, y=50)

    page10_label1 = Label(page10, text='AVAILABLE ', font=('Arial', 16, 'bold') , bg="light green")
    page10_label1.place(x=270, y=120)

    page10_label2 = Label(page10, text='NOT AVAILABLE ', font=('Arial', 16, 'bold') , bg="light green")
    page10_label2.place(x=890, y=120)

    back6 = Button(page10, text="Back", fg="Black", 
                                bg="Red", command=lambda: show_frame(page1)) 
    back6.place(x = 50, y = 500 ,width = 100 , height = 25)
    
    scrollbox1 = Listbox(page10 , font= ('Helvetica 12 bold'))
    scrollbox1.place(x=170, y=205 , width = 320 , height = 200)

    aval_num = 0
    notaval_num = 0
    
    for i in range(2, sh.max_row+1):
        stat_avl1 = sh.cell(row=i, column=7)
        if(stat_avl1.value == "Available"):
            serial_scrolla = sh.cell(row=i, column=1)
            scrollbox1.insert('end' , serial_scrolla.value)
            aval_num = aval_num + 1

    page10_num = Label(page10, text='Available ' + str(aval_num), font=('Arial', 14, 'bold') , bg="light green")
    page10_num.place(x=270, y=160)

    scrollbox2 = Listbox(page10 , font= ('Helvetica 12 bold'))
    scrollbox2.place(x=803, y=205 , width = 320 , height = 200)

    for i in range(2, sh.max_row+1):
        stat_avl2 = sh.cell(row=i, column=7)
        if(stat_avl2.value == "Not Available"):
            serial_scrolln = sh.cell(row=i, column=1)
            scrollbox2.insert('end' , serial_scrolln.value)
            notaval_num = notaval_num + 1

    page10_num1 = Label(page10, text='Not Available ' + str(notaval_num), font=('Arial', 14, 'bold') , bg="light green")
    page10_num1.place(x=900, y=160)

    

#-----------Main program    

page1 = Frame(window)
page2 = Frame(window)
page3 = Frame(window)
page4 = Frame(window)
page5 = Frame(window)
page6 = Frame(window)
page7 = Frame(window)
page8 = Frame(window)
page9 = Frame(window)
page10 = Frame(window)
page11 = Frame(window)
page12 = Frame(window)

for frame in (page1, page2, page3, page4, page5, page6, page7, page8, page9, page10, page11, page12):
    frame.grid(row=0, column=0, sticky='nsew')

def show_frame(frame):
    frame.tkraise()

show_frame(page1)

#----------------------------page1

page1.config(background='white')
pag1_label = Label(page1, text='WELCOME', font=('Arial', 20, 'bold'), bg="white")
pag1_label.place(x=600, y=100)
  
    # set the title of GUI window

submit = Button(page1, text="Take", fg="Black", 
                            bg="Red", command=lambda: show_frame(page4))
submit.place(x=195, y=350 , width = 220 , height = 30)

ret = Button(page1, text="Return", fg="Black", 
                            bg="Red", command=lambda: show_frame(page6)) 
ret.place(x = 550, y = 350 ,width = 220 , height = 30)

new = Button(page1, text="Add New Items", fg="Black", 
                            bg="Red", command=lambda: show_frame(page3)) 
new.place(x = 900, y = 350 ,width = 220 , height = 30)

item_l = Button(page1, text="Item List", fg="Black", 
                            bg="Red", command=scroll_list) 
item_l.place(x = 550, y = 400 ,width = 220 , height = 30)

#----------------------------page2

page2.config(background='light green')
pag2_status = Label(page2, text='ITEM AVAILABLE', font=('Arial', 20, 'bold') , bg="light green")
pag2_status.place(x=590, y=20)
pag2_label = Label(page2, text='FORM', font=('Arial', 20, 'bold') , bg="light green")
pag2_label.place(x=640, y=90)
heading = Label(page2, text="Form", bg="light green") 
  
    # create a Name label 
name = Label(page2, text="Name", font=('Arial' , 15, 'bold') , bg="light green")
name.place(x=50, y=180)
  
    # create a Course label 
course = Label(page2, text="Email", font=('Arial' , 15, 'bold') , bg="light green")
course.place(x=50, y=230)
  
    # create a Semester label 
sem = Label(page2, text="Expected Return Date", font=('Arial' , 15, 'bold') , bg="light green")
sem.place(x=50, y=280)
  
    # create a Form No. lable
form_no = Label(page2, text="Phone Number.", font=('Arial' , 15, 'bold') , bg="light green")
form_no.place(x=50, y=330)
 
  
    # create a text entry box 
    # for typing the information 
name_field = Entry(page2)
name_field.place(x=230, y=185 , width = 220 , height = 25)
course_field = Entry(page2)
course_field.place(x=230, y=235 , width = 220 , height = 25)
sem_field = Entry(page2)
sem_field.place(x=275, y=285 , width = 220 , height = 25)
form_no_field = Entry(page2)
form_no_field.place(x=230, y=335 , width = 220 , height = 25)
  
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
name_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
course_field.bind("<Return>", focus2) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
sem_field.bind("<Return>", focus3) 
  
    # whenever the enter key is pressed  
  
    # call excel function 
excel() 
  
    # create a Submit Button and place into the root window 
submit = Button(page2, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
submit.place(x = 800, y = 500 , width = 100 , height = 25)

back = Button(page2, text="Back", fg="Black", 
                            bg="Red", command=lambda: show_frame(page4)) 
back.place(x = 50, y = 500 ,width = 100 , height = 25)

#----------------------------page3

page3.config(background='light green')
pag3_label = Label(page3, text='NEW ITEMS', font=('Arial', 20, 'bold') , bg="light green")
pag3_label.place(x=600, y=50)
heading = Label(page3, text="Form", bg="light green") 
  
    # create a Name label 
name = Label(page3, text="Serial Number", font=('Arial' , 15, 'bold') , bg="light green")
name.place(x=50, y=100)

Serial_field = Entry(page3)
Serial_field.place(x=230, y=105 , width = 220 , height = 25)

mlfb = Label(page3, text="MLFB Number", font=('Arial' , 15, 'bold') , bg="light green")
mlfb.place(x=50, y=150)

Serial_field4 = Entry(page3)
Serial_field4.place(x=230, y=155 , width = 220 , height = 25)

excel() 
  
    # create a Submit Button and place into the root window 
submit1 = Button(page3, text="Enter", fg="Black", 
                            bg="Red", command=insert_page3) 
submit1.place(x = 900, y = 400 , width = 100 , height = 25)

back1 = Button(page3, text="Back", fg="Black", 
                            bg="Red", command=lambda: show_frame(page1)) 
back1.place(x = 50, y = 500 ,width = 100 , height = 25)


#-----------------------------page4

page4.config(background='light green')
page4_label = Label(page4, text='CHECK AVAILABILITY', font=('Arial', 20, 'bold') , bg="light green")
page4_label.place(x=570, y=50)
heading = Label(page4, text="Form", bg="light green") 
  
    # create a Name label 
name = Label(page4, text="Serial Number", font=('Arial' , 15, 'bold') , bg="light green")
name.place(x=50, y=100)

sh = wb.active
def Scankey(event):

    sh = wb.active
    list.clear()

    for i in range(2, sh.max_row+1):
        serial_list = sh.cell(row=i, column=1)
        list.append(serial_list.value)
            
    val = event.widget.get()
    #print(val)
    if val == '':
        data = list
    else:
        data = []
        for item in list:
            if val.lower() in item.lower():
                data.append(item)
        
    Update(data)
    window.bind('<Double-1>', click)

def Update(data):
    listbox.delete(0, 'end')
    for item in data:
        listbox.insert('end', item)
    #Serail_field1 = listbox.select_set(0)
    
    #textbox.pack(pady=20)
    #Serial_field1 = variable.get()

def click(event):
    for i in listbox.curselection():
        #print(listbox.get(i))
        Serial_field1.delete(0, END)
        Serial_field1.insert('end' , listbox.get(i))
list = []
# iterate through excel and display data





variable = StringVar(page4)
variable.set(list) # default value

#w = OptionMenu(page4, variable, *list)
#w.pack()


Serial_field1 = Entry(page4)
Serial_field1.place(x=230, y=105 , width = 200 , height = 25)
#Serial_field1.pack()
Serial_field1.bind('<KeyRelease>', Scankey)

listbox = Listbox(page4)
listbox.place(x=230, y=130 , width = 200 , height = 125)
#listbox.pack()
Update(list)



submit2 = Button(page4, text="Check", fg="Black", 
                            bg="Red", command=check) 
submit2.place(x = 900, y = 400 , width = 100 , height = 25)

back2 = Button(page4, text="Back", fg="Black", 
                            bg="Red", command=lambda: show_frame(page1)) 
back2.place(x = 50, y = 500 ,width = 100 , height = 25)


#------------------------------page5

check()

#-----------------------------page6

page6.config(background='light green')
page6_label = Label(page6, text='RETURN ITEM', font=('Arial', 20, 'bold') , bg="light green")
page6_label.place(x=570, y=50)
heading = Label(page6, text="Form", bg="light green") 
  
    # create a Name label 
name5 = Label(page6, text="Serial Number", font=('Arial' , 15, 'bold') , bg="light green")
name5.place(x=50, y=100)

def Scankey1(event):
    sh = wb.active
    list1.clear()

    for i in range(2, sh.max_row+1):
        stat_av = sh.cell(row=i, column=7)
        #serial_list1 = sh.cell(row=i, column=1)
        #list1.append(serial_list1.value)
        if(stat_av.value == "Not Available"):
            
            serial_list1 = sh.cell(row=i, column=1)
            list1.append(serial_list1.value)
            
    val1 = event.widget.get()
    #print(val)
    if val1 == '':
        data = list1
    else:
        data = []
        for item in list1:
            if val1.lower() in item.lower():
                data.append(item)

    Update1(data)
    window.bind('<Double-1>', click1)

def Update1(data):
    listbox1.delete(0, 'end')
    for item in data:
        listbox1.insert('end', item)
    #Serail_field1 = listbox.select_set(0)
    
    #textbox.pack(pady=20)
    #Serial_field1 = variable.get()

def click1(event):
    for i in listbox1.curselection():
        #print(listbox.get(i))
        Serial_field2.delete(0, END)
        Serial_field2.insert('end' , listbox1.get(i))
list1 = []
# iterate through excel and display data




variable1 = StringVar(page6)
variable1.set(list1) # default value

#w = OptionMenu(page4, variable, *list)
#w.pack()


Serial_field2 = Entry(page6)
Serial_field2.place(x=230, y=105 , width = 200 , height = 25)
#Serial_field1.pack()
Serial_field2.bind('<KeyRelease>', Scankey1)

listbox1 = Listbox(page6)
listbox1.place(x=230, y=130 , width = 200 , height = 125)
#listbox.pack()
Update1(list1)

submit3 = Button(page6, text="Confirm Return", fg="Black", 
                            bg="Red", command=returnitem) 
submit3.place(x = 900, y = 400 , width = 100 , height = 25)

back2 = Button(page6, text="Back", fg="Black", 
                            bg="Red", command=lambda: show_frame(page1)) 
back2.place(x = 50, y = 500 ,width = 100 , height = 25)

#----------------------------page7



#----------------------------page8

page8.config(background='light green')
page8_label = Label(page8, text='Submitted Successfully ', font=('Arial', 20, 'bold') , bg="light green")
page8_label.place(x=570, y=200)

back4 = Button(page8, text="Back", fg="Black", 
                            bg="Red", command=lambda: show_frame(page4)) 
back4.place(x = 50, y = 500 ,width = 100 , height = 25)

#----------------------------page9

page9.config(background='light green')
page9_label = Label(page9, text='Item Successfully Added ', font=('Arial', 20, 'bold') , bg="light green")
page9_label.place(x=570, y=200)

back5 = Button(page9, text="Back", fg="Black", 
                            bg="Red", command=lambda: show_frame(page1)) 
back5.place(x = 50, y = 500 ,width = 100 , height = 25)


#----------------------------page10

#----------------------------page11

    # start the GUI 
window.mainloop()
